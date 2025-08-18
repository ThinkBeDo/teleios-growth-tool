#!/usr/bin/env python3
"""
Create a proper sample County Excel file for testing the Teleios Growth Tool
"""
from openpyxl import Workbook
import sys

def create_sample_county_file(county_name="Sample"):
    """Create a properly formatted County Excel file"""
    
    filename = f"{county_name}.xlsx"
    
    # Create workbook and get active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "County Trend"
    
    # Add title rows (optional, but common in real files)
    ws['A1'] = f"{county_name} County Health Statistics"
    ws['A2'] = "Medicare and Hospice Data"
    
    # Add headers at row 9 (as expected by the extraction code)
    ws['A9'] = 'Row'
    ws['B9'] = 'Year'
    ws['C9'] = 'Medicare Enrollment'
    ws['D9'] = 'Medicare %'
    ws['E9'] = 'Resident Deaths'
    ws['F9'] = 'Death Rate'
    ws['G9'] = 'Hospice Deaths'
    ws['H9'] = 'Hospice %'
    ws['I9'] = 'Patients Served'
    ws['J9'] = 'Service Rate'
    
    # Add 15 rows of sample data starting at row 10
    base_year = 2009
    base_medicare = 15000
    base_deaths = 800
    base_hospice_deaths = 250
    base_patients = 1200
    
    for i in range(15):
        row = 10 + i
        year = base_year + i
        
        # Generate realistic looking data with some variation
        medicare = base_medicare + (i * 250) + (i % 3 * 50)
        deaths = base_deaths + (i * 15) + (i % 2 * 10)
        hospice_deaths = base_hospice_deaths + (i * 8) + (i % 3 * 5)
        patients = base_patients + (i * 35) + (i % 2 * 20)
        
        # Fill in the data
        ws[f'A{row}'] = i + 1  # Row number
        ws[f'B{row}'] = year
        ws[f'C{row}'] = medicare
        ws[f'D{row}'] = round(medicare / 50000 * 100, 1)  # Percentage
        ws[f'E{row}'] = deaths
        ws[f'F{row}'] = round(deaths / medicare * 1000, 1)  # Rate per 1000
        ws[f'G{row}'] = hospice_deaths
        ws[f'H{row}'] = round(hospice_deaths / deaths * 100, 1)  # Percentage
        ws[f'I{row}'] = patients
        ws[f'J{row}'] = round(patients / medicare * 1000, 1)  # Rate per 1000
    
    # Add a note at row 26 (after the data)
    ws['A26'] = "Note: This is sample data for testing purposes only"
    
    # Save the workbook
    wb.save(filename)
    print(f"✅ Created sample county file: {filename}")
    print(f"   - Contains 'County Trend' sheet")
    print(f"   - Headers at row 9")
    print(f"   - 15 rows of data (rows 10-24)")
    print(f"   - Years: {base_year} to {base_year + 14}")
    print(f"\nThis file can be used to test the Teleios Growth Tool")
    
    return filename

def main():
    """Main function"""
    if len(sys.argv) > 1:
        county_name = sys.argv[1]
    else:
        county_name = "Sample"
    
    print(f"Creating sample Excel file for {county_name} County...")
    create_sample_county_file(county_name)
    
    print("\nTo create files for other counties, run:")
    print("  python create_sample_county.py CountyName")
    print("\nExample:")
    print("  python create_sample_county.py Stokes")
    print("  python create_sample_county.py Durham")

if __name__ == "__main__":
    main()