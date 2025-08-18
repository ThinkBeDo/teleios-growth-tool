from flask import Flask, request, render_template, send_file, flash, redirect, url_for, jsonify
import os
import io
from werkzeug.utils import secure_filename
from utils.excel_processor import process_county_files, extract_county_data
from utils.debug_logger import debug_logger
import tempfile
import zipfile
from openpyxl import load_workbook
import json

app = Flask(__name__)
app.secret_key = 'teleios-growth-tool-secret-key'

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    try:
        # Check if main workbook is provided
        if 'main_workbook' not in request.files:
            flash('No main workbook file selected')
            return redirect(request.url)
        
        main_file = request.files['main_workbook']
        if main_file.filename == '':
            flash('No main workbook file selected')
            return redirect(request.url)
        
        # Check if county files are provided
        county_files = request.files.getlist('county_files')
        if not county_files or all(f.filename == '' for f in county_files):
            flash('No county files selected')
            return redirect(request.url)
        
        # Validate file types
        if not allowed_file(main_file.filename):
            flash('Main workbook must be an Excel file (.xlsx or .xls)')
            return redirect(request.url)
        
        valid_county_files = []
        for file in county_files:
            if file.filename != '' and allowed_file(file.filename):
                valid_county_files.append(file)
        
        if not valid_county_files:
            flash('No valid county Excel files found')
            return redirect(request.url)
        
        # Process the files
        result_file = process_county_files(main_file, valid_county_files)
        
        if result_file:
            # Return the processed file
            return send_file(
                result_file,
                as_attachment=True,
                download_name='updated_growth_tool.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('Error processing files. Some county files may be corrupted or not in proper Excel format. Please ensure all files are valid .xlsx files exported from Excel.')
            return redirect(url_for('index'))
            
    except Exception as e:
        flash(f'Error processing files: {str(e)}')
        return redirect(url_for('index'))

@app.route('/health')
def health_check():
    return {'status': 'healthy', 'service': 'Teleios Growth Tool'}

# Debug endpoints
@app.route('/debug/file-info', methods=['POST'])
def debug_file_info():
    """Analyze file structure without processing"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Reset debug logger
        debug_logger.reset_trace()
        
        # Analyze file
        file_info = debug_logger.log_file_info(file)
        
        # Try to load as Excel and get sheet info
        sheet_info = {}
        try:
            file.seek(0)
            wb = load_workbook(file, data_only=True)
            sheet_info = {
                'sheet_names': wb.sheetnames,
                'sheet_count': len(wb.sheetnames),
                'county_trend_found': any('trend' in s.lower() and 'county' in s.lower() for s in wb.sheetnames)
            }
            
            # If County Trend sheet exists, get its info
            for sheet_name in wb.sheetnames:
                if 'trend' in sheet_name.lower() and 'county' in sheet_name.lower():
                    sheet = wb[sheet_name]
                    sheet_info['county_trend_info'] = {
                        'name': sheet_name,
                        'max_row': sheet.max_row,
                        'max_column': sheet.max_column,
                        'row_10_data': {
                            'B': sheet['B10'].value if sheet.max_row >= 10 else None,
                            'C': sheet['C10'].value if sheet.max_row >= 10 else None,
                            'E': sheet['E10'].value if sheet.max_row >= 10 else None,
                            'G': sheet['G10'].value if sheet.max_row >= 10 else None,
                            'I': sheet['I10'].value if sheet.max_row >= 10 else None,
                        }
                    }
                    break
        except Exception as e:
            sheet_info['error'] = str(e)
        
        return jsonify({
            'file_info': file_info,
            'sheet_info': sheet_info,
            'trace': debug_logger.get_full_trace()
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'trace': debug_logger.get_full_trace()}), 500

@app.route('/debug/extraction-test', methods=['POST'])
def debug_extraction_test():
    """Test extraction with detailed logging"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Run extraction with debug logging
        extracted_data = extract_county_data(file)
        
        # Get the full trace
        trace = debug_logger.get_full_trace()
        
        return jsonify({
            'filename': file.filename,
            'extraction_successful': extracted_data is not None,
            'rows_extracted': len(extracted_data) if extracted_data else 0,
            'data_sample': extracted_data[:3] if extracted_data else None,
            'trace': trace,
            'expected_15_rows': len(extracted_data) == 15 if extracted_data else False
        })
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'trace': debug_logger.get_full_trace()
        }), 500

@app.route('/debug/row-count', methods=['POST'])
def debug_row_count():
    """Count rows in each section of County Trend sheet"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        wb = load_workbook(file, data_only=True)
        
        # Find County Trend sheet
        trend_sheet = None
        trend_sheet_name = None
        for sheet_name in wb.sheetnames:
            if 'trend' in sheet_name.lower() and 'county' in sheet_name.lower():
                trend_sheet = wb[sheet_name]
                trend_sheet_name = sheet_name
                break
        
        if not trend_sheet:
            return jsonify({'error': 'No County Trend sheet found'}), 404
        
        # Analyze data sections
        sections = []
        current_section = None
        empty_row_count = 0
        
        for row in range(1, trend_sheet.max_row + 1):
            year_val = trend_sheet[f'B{row}'].value
            medicare_val = trend_sheet[f'C{row}'].value
            
            if year_val is not None and medicare_val is not None:
                # Data row found
                if current_section is None or empty_row_count > 0:
                    # Start new section
                    if current_section:
                        sections.append(current_section)
                    current_section = {
                        'start_row': row,
                        'end_row': row,
                        'row_count': 1,
                        'first_year': year_val,
                        'last_year': year_val
                    }
                    empty_row_count = 0
                else:
                    # Continue current section
                    current_section['end_row'] = row
                    current_section['row_count'] += 1
                    current_section['last_year'] = year_val
            else:
                # Empty row
                empty_row_count += 1
        
        # Add last section if exists
        if current_section:
            sections.append(current_section)
        
        # Check our extraction logic
        extraction_test = {
            'would_extract_rows_10_to': None,
            'would_extract_count': 0
        }
        
        for row in range(10, trend_sheet.max_row + 1):
            year_val = trend_sheet[f'B{row}'].value
            medicare_val = trend_sheet[f'C{row}'].value
            
            if year_val is None or year_val == '' or medicare_val is None or medicare_val == '':
                extraction_test['would_extract_rows_10_to'] = row - 1
                extraction_test['would_extract_count'] = row - 10
                break
        
        if extraction_test['would_extract_rows_10_to'] is None:
            extraction_test['would_extract_rows_10_to'] = trend_sheet.max_row
            extraction_test['would_extract_count'] = trend_sheet.max_row - 9
        
        return jsonify({
            'filename': file.filename,
            'sheet_name': trend_sheet_name,
            'max_row': trend_sheet.max_row,
            'sections_found': len(sections),
            'sections': sections,
            'extraction_test': extraction_test,
            'issue_detected': extraction_test['would_extract_count'] != 15
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
