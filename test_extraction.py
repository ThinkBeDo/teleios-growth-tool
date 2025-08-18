#!/usr/bin/env python3
"""
Test script for debugging Teleios Growth Tool extraction issues
"""
import os
import sys
import io
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import zipfile

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from utils.excel_processor import extract_county_data
from utils.debug_logger import debug_logger

def create_test_excel_files():
    """Create test Excel files with various scenarios"""
    test_files = []
    
    # 1. Create a proper Excel file with 15 rows (expected behavior)
    print("Creating test file: Proper_County.xlsx (15 rows)")
    wb = Workbook()
    ws = wb.active
    ws.title = "County Trend"
    
    # Add headers at row 9
    ws['B9'] = 'Year'
    ws['C9'] = 'Medicare Enrollment'
    ws['E9'] = 'Resident Deaths'
    ws['G9'] = 'Hospice Deaths'
    ws['I9'] = 'Patients Served'
    
    # Add 15 rows of data starting at row 10
    for i in range(15):
        row = 10 + i
        ws[f'B{row}'] = 2009 + i  # Years 2009-2023
        ws[f'C{row}'] = 10000 + (i * 100)  # Medicare enrollment
        ws[f'E{row}'] = 500 + (i * 10)  # Resident deaths
        ws[f'G{row}'] = 200 + (i * 5)  # Hospice deaths
        ws[f'I{row}'] = 1000 + (i * 20)  # Patients served
    
    wb.save('test_files/Proper_County.xlsx')
    test_files.append('test_files/Proper_County.xlsx')
    
    # 2. Create Excel with 3 sections (45 rows total - the bug scenario)
    print("Creating test file: Triple_Section_County.xlsx (45 rows in 3 sections)")
    wb = Workbook()
    ws = wb.active
    ws.title = "County Trend"
    
    # Headers
    ws['B9'] = 'Year'
    ws['C9'] = 'Medicare Enrollment'
    ws['E9'] = 'Resident Deaths'
    ws['G9'] = 'Hospice Deaths'
    ws['I9'] = 'Patients Served'
    
    # First section: 15 rows (10-24)
    for i in range(15):
        row = 10 + i
        ws[f'B{row}'] = 2009 + i
        ws[f'C{row}'] = 10000 + (i * 100)
        ws[f'E{row}'] = 500 + (i * 10)
        ws[f'G{row}'] = 200 + (i * 5)
        ws[f'I{row}'] = 1000 + (i * 20)
    
    # Empty rows at 25-26
    
    # Second section: 15 rows (27-41) - maybe projections
    ws['B26'] = 'Projections'
    for i in range(15):
        row = 27 + i
        ws[f'B{row}'] = 2024 + i
        ws[f'C{row}'] = 12000 + (i * 100)
        ws[f'E{row}'] = 600 + (i * 10)
        ws[f'G{row}'] = 250 + (i * 5)
        ws[f'I{row}'] = 1200 + (i * 20)
    
    # Empty rows at 42-43
    
    # Third section: 15 rows (44-58) - maybe adjusted data
    ws['B43'] = 'Adjusted'
    for i in range(15):
        row = 44 + i
        ws[f'B{row}'] = 2009 + i
        ws[f'C{row}'] = 11000 + (i * 100)
        ws[f'E{row}'] = 550 + (i * 10)
        ws[f'G{row}'] = 225 + (i * 5)
        ws[f'I{row}'] = 1100 + (i * 20)
    
    wb.save('test_files/Triple_Section_County.xlsx')
    test_files.append('test_files/Triple_Section_County.xlsx')
    
    # 3. Create a corrupted "Excel" file (not a valid zip)
    print("Creating test file: FakeCounty_NotZip.xlsx (corrupted file)")
    with open('test_files/FakeCounty_NotZip.xlsx', 'wb') as f:
        f.write(b'This is not a valid Excel file or zip archive')
    test_files.append('test_files/FakeCounty_NotZip.xlsx')
    
    # 4. Create an old-format Excel file simulation (XLS header)
    print("Creating test file: OldFormat_County.xls (XLS format simulation)")
    with open('test_files/OldFormat_County.xls', 'wb') as f:
        # XLS file signature
        f.write(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1')
        f.write(b'Rest of file would be OLE2 compound document format')
    test_files.append('test_files/OldFormat_County.xls')
    
    return test_files

def test_file_extraction(filepath):
    """Test extraction on a single file"""
    print(f"\n{'='*60}")
    print(f"Testing: {filepath}")
    print('='*60)
    
    # Reset debug logger
    debug_logger.reset_trace()
    
    # Create file-like object
    try:
        with open(filepath, 'rb') as f:
            file_obj = io.BytesIO(f.read())
            file_obj.filename = os.path.basename(filepath)
            file_obj.seek(0)
            
            # Test file info
            print("\n1. Analyzing file structure...")
            file_info = debug_logger.log_file_info(file_obj)
            print(f"   - File size: {file_info.get('file_size', 'Unknown')} bytes")
            print(f"   - Is valid zip: {file_info.get('is_zip_file', False)}")
            print(f"   - File type: {file_info.get('file_type', 'Unknown')}")
            
            if file_info.get('zip_test'):
                print(f"   - Zip test error: {file_info['zip_test']}")
            
            # Try extraction
            print("\n2. Attempting data extraction...")
            file_obj.seek(0)
            extracted_data = extract_county_data(file_obj)
            
            if extracted_data:
                print(f"   ✓ Extraction successful!")
                print(f"   - Rows extracted: {len(extracted_data)}")
                print(f"   - Expected 15 rows: {'YES' if len(extracted_data) == 15 else 'NO - ISSUE DETECTED!'}")
                
                if len(extracted_data) > 0:
                    print(f"   - First row year: {extracted_data[0]['year']}")
                    print(f"   - Last row year: {extracted_data[-1]['year']}")
            else:
                print(f"   ✗ Extraction failed!")
                
            # Get trace summary
            trace = debug_logger.get_full_trace()
            print(f"\n3. Debug trace steps: {trace['trace_count']}")
            
            # Find summary in trace
            for step in trace['trace']:
                if step['step'] == 'summary':
                    summary = step['data']
                    print(f"   - Total rows checked: {summary['total_rows_checked']}")
                    print(f"   - Rows extracted: {summary['rows_extracted']}")
                    print(f"   - Stop reason: {summary['stop_reason']}")
                    
            # Save trace if there's an issue
            if not extracted_data or (extracted_data and len(extracted_data) != 15):
                trace_file = debug_logger.save_trace_to_file(f"test_{os.path.basename(filepath)}")
                print(f"\n   ⚠ Issue detected - trace saved to: {trace_file}")
                
    except Exception as e:
        print(f"   ✗ Error during testing: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Main test execution"""
    print("Teleios Growth Tool - Extraction Debug Test Suite")
    print("="*60)
    
    # Create test directory
    os.makedirs('test_files', exist_ok=True)
    os.makedirs('debug_traces', exist_ok=True)
    
    # Create test files
    print("\nCreating test Excel files...")
    test_files = create_test_excel_files()
    
    # Test each file
    print("\n\nRunning extraction tests...")
    for filepath in test_files:
        test_file_extraction(filepath)
    
    # Summary
    print(f"\n{'='*60}")
    print("TEST SUMMARY")
    print('='*60)
    print("Expected behaviors:")
    print("1. Proper_County.xlsx - Should extract exactly 15 rows ✓")
    print("2. Triple_Section_County.xlsx - Should extract only 15 rows (not 45) ✓")
    print("3. FakeCounty_NotZip.xlsx - Should fail with 'not a valid Excel file' error ✓")
    print("4. OldFormat_County.xls - Should fail or handle XLS format appropriately ✓")
    
    print("\nCheck debug_traces/ folder for detailed extraction traces.")
    print("\nTo test with real files (Stokes.xlsx, etc.), place them in test_files/ and run:")
    print("  python test_extraction.py test_files/Stokes.xlsx")

if __name__ == "__main__":
    # Check if specific file was provided as argument
    if len(sys.argv) > 1:
        test_file_extraction(sys.argv[1])
    else:
        main()