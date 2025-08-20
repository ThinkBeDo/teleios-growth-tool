import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_PERCENTAGE_00
import io
import tempfile
import os
from datetime import datetime
from .debug_logger import debug_logger
import zipfile

def process_county_files(main_workbook_file, county_files):
    """
    Process county Excel files and update the main workbook's Raw sheet.
    
    Field Mapping (12 data columns WITH GAPS for calculated fields):
    - Column C (Medicare Enrollment) → Raw Column E (Medicare Beneficiaries)
    - Column E (Resident Deaths) → Raw Column G (Medicare Deaths) [Skip F - calculated]
    - Column G (Hospice Deaths) → Raw Column I (Hospice Deaths) [Skip H - calculated]
    - Column I (Patients Served) → Raw Column K (Hospice Unduplicated Beneficiaries)
    - Column H (Hospice Penetration) → Raw Column J (Hospice Penetration)
    - Column J (Days per Patient/ALOS) → Raw Column L (Days per Patient)
    - Column K (Patient Days) → Raw Column M (Patient Days)
    - Column L (Average Daily Census) → Raw Column N (Average Daily Census)
    - Column M (% GIP Days) → Raw Column O (% GIP Days)
    - Column N (Average GIP Census) → Raw Column P (Average GIP Census)
    - Column O (GIP Patients) → Raw Column Q (GIP Patients)
    - Column P (Payments per Patient) → Raw Column R (Payments per Patient)
    
    NOTE: Columns F and H in Raw sheet are CALCULATED fields - DO NOT OVERWRITE
    """
    try:
        # Load main workbook
        main_wb = load_workbook(main_workbook_file, data_only=False)
        
        # Find the Raw sheet
        if 'Raw' not in main_wb.sheetnames:
            raise ValueError("Main workbook does not contain a 'Raw' sheet")
        
        raw_sheet = main_wb['Raw']
        
        # Find the next empty row in Raw sheet
        next_row = find_next_empty_row(raw_sheet)
        
        # Process each county file
        for county_file in county_files:
            county_data = extract_county_data(county_file)
            if county_data:
                # Add data to Raw sheet
                next_row = add_county_data_to_raw(raw_sheet, county_data, next_row)
        
        # CRITICAL FIX: Standardize KEY columns to ensure XLOOKUP works
        standardize_key_columns(main_wb)
        
        # CRITICAL FIX: APPEND new counties to Counties sheet (preserves existing data)
        # This ensures new counties are added WITHOUT deleting existing county data
        if 'Counties' in main_wb.sheetnames:
            new_rows_added = append_new_counties_to_sheet(main_wb)
            
            # Get the max row from Raw sheet for XLOOKUP range
            raw_max_row = raw_sheet.max_row
            
            # Restore XLOOKUP formulas for ALL rows (both existing and new)
            restore_xlookup_formulas(main_wb['Counties'], raw_max_row)
            
            # Apply formatting to all rows
            apply_counties_sheet_formatting(main_wb['Counties'])
        
        # Save to memory buffer
        output_buffer = io.BytesIO()
        main_wb.save(output_buffer)
        output_buffer.seek(0)
        
        return output_buffer
        
    except Exception as e:
        print(f"Error processing files: {e}")
        return None

def append_new_counties_to_sheet(workbook):
    """
    APPEND new counties from Raw sheet to Counties sheet WITHOUT deleting existing data.
    This preserves all existing county data and only adds new counties at the end.
    """
    try:
        raw_sheet = workbook['Raw']
        counties_sheet = workbook['Counties']
        
        # Get existing counties to avoid duplicates
        existing_counties = set()
        for row in range(2, counties_sheet.max_row + 1):
            county = counties_sheet[f'D{row}'].value
            year = counties_sheet[f'A{row}'].value
            if county and year:
                existing_counties.add((county, year))
        
        # Find the next empty row in Counties sheet (where to append)
        next_counties_row = counties_sheet.max_row + 1
        if counties_sheet.max_row == 1:  # Only headers exist
            next_counties_row = 2
        else:
            # Find actual last row with data
            for row in range(counties_sheet.max_row, 1, -1):
                if counties_sheet[f'A{row}'].value is not None:
                    next_counties_row = row + 1
                    break
        
        # Get new counties from Raw sheet
        new_counties_added = 0
        row = 2  # Start from row 2 in Raw sheet (skip header)
        
        while raw_sheet[f'A{row}'].value is not None:
            county = raw_sheet[f'A{row}'].value
            state = raw_sheet[f'B{row}'].value
            year = raw_sheet[f'C{row}'].value
            
            # Only add if this county-year combination doesn't already exist
            if county and state and year and (county, year) not in existing_counties:
                # Map data from Raw sheet to Counties sheet structure
                counties_sheet[f'A{next_counties_row}'] = year  # Year
                counties_sheet[f'B{next_counties_row}'] = year  # Year (duplicate)
                counties_sheet[f'C{next_counties_row}'] = state  # State
                counties_sheet[f'D{next_counties_row}'] = county  # County
                
                # Key field
                key_value = f"{county}{year}"
                counties_sheet[f'E{next_counties_row}'] = key_value
                
                # Don't add formulas here - they will be added by restore_xlookup_formulas
                # This ensures all rows use consistent XLOOKUP formulas
                
                next_counties_row += 1
                new_counties_added += 1
                existing_counties.add((county, year))
            
            row += 1
        
        if new_counties_added > 0:
            print(f"Successfully appended {new_counties_added} new county rows to Counties sheet")
        else:
            print("No new counties to add (all already exist in Counties sheet)")
        
        return new_counties_added
        
    except Exception as e:
        print(f"Error appending to Counties sheet: {e}")
        return 0

def rebuild_counties_sheet_from_raw(workbook):
    """
    DEPRECATED - Use append_new_counties_to_sheet() instead.
    This function rebuilds from scratch and loses existing data.
    Kept for backwards compatibility but should not be used.
    """
    try:
        raw_sheet = workbook['Raw']
        counties_sheet = workbook['Counties']
        
        # Get all unique counties from Raw sheet
        counties_data = []
        row = 2  # Start from row 2 (skip header)
        
        while raw_sheet[f'A{row}'].value is not None:
            county = raw_sheet[f'A{row}'].value
            state = raw_sheet[f'B{row}'].value
            year = raw_sheet[f'C{row}'].value
            
            if county and state and year:
                counties_data.append({
                    'county': county,
                    'state': state, 
                    'year': year,
                    'raw_row': row
                })
            row += 1
        
        # Clear existing Counties sheet data (keep headers)
        max_row = counties_sheet.max_row
        if max_row > 1:
            counties_sheet.delete_rows(2, max_row - 1)
        
        # Rebuild Counties sheet with all counties from Raw
        current_counties_row = 2
        
        for data in counties_data:
            # Map data from Raw sheet to Counties sheet structure
            counties_sheet[f'A{current_counties_row}'] = data['year']  # Year
            counties_sheet[f'B{current_counties_row}'] = data['year']  # Year (duplicate)
            counties_sheet[f'C{current_counties_row}'] = data['state']  # State
            counties_sheet[f'D{current_counties_row}'] = data['county']  # County
            
            # Key field - FIXED: Use direct concatenation instead of CONCAT function
            # Format: CountyYear (e.g., "Caswell2010")
            key_value = f"{data['county']}{data['year']}"
            counties_sheet[f'E{current_counties_row}'] = key_value
            
            # Copy data from Raw sheet using direct cell references - WITH GAPS for calculated fields
            # Medicare Beneficiaries (Raw column E)
            counties_sheet[f'H{current_counties_row}'] = f'=Raw!E{data["raw_row"]}'
            
            # Medicare Deaths (Raw column G - NOT F which is calculated) 
            counties_sheet[f'I{current_counties_row}'] = f'=Raw!G{data["raw_row"]}'
            
            # Hospice Unduplicated Beneficiaries (Raw column K)
            counties_sheet[f'J{current_counties_row}'] = f'=Raw!K{data["raw_row"]}'
            
            # Hospice Deaths (Raw column I)
            counties_sheet[f'K{current_counties_row}'] = f'=Raw!I{data["raw_row"]}'
            
            # Hospice Penetration (Raw column J) → Counties column Z
            counties_sheet[f'Z{current_counties_row}'] = f'=Raw!J{data["raw_row"]}'
            
            # Average Daily Census (Raw column N) → Counties column AB
            counties_sheet[f'AB{current_counties_row}'] = f'=Raw!N{data["raw_row"]}'
            
            # Patient Days (Raw column M) → Counties column AC
            counties_sheet[f'AC{current_counties_row}'] = f'=Raw!M{data["raw_row"]}'
            
            # Days per Patient/ALOS (Raw column L) → Counties column AD
            counties_sheet[f'AD{current_counties_row}'] = f'=Raw!L{data["raw_row"]}'
            
            # % GIP Days (Raw column O) → Counties column AE
            counties_sheet[f'AE{current_counties_row}'] = f'=Raw!O{data["raw_row"]}'
            
            # Average GIP Census (Raw column P) → Counties column AF
            counties_sheet[f'AF{current_counties_row}'] = f'=Raw!P{data["raw_row"]}'
            
            # GIP Patients (Raw column Q) → Counties column AG
            counties_sheet[f'AG{current_counties_row}'] = f'=Raw!Q{data["raw_row"]}'
            
            # Payments per Patient (Raw column R) → Counties column AH
            counties_sheet[f'AH{current_counties_row}'] = f'=Raw!R{data["raw_row"]}'
            
            current_counties_row += 1
        
        # Apply consistent formatting to the Counties sheet
        apply_counties_sheet_formatting(counties_sheet)
        
        print(f"Successfully rebuilt Counties sheet with {len(counties_data)} rows from Raw sheet")
        
    except Exception as e:
        print(f"Error rebuilding Counties sheet: {e}")
        # Don't fail the entire process if Counties rebuild fails
        pass

def standardize_key_columns(workbook):
    """
    Standardize KEY columns in both Raw and Counties sheets to use CountyYear format.
    This ensures XLOOKUP formulas can find matches between sheets.
    
    Args:
        workbook: The workbook containing Raw and Counties sheets
    """
    try:
        # Check if sheets exist
        if 'Raw' not in workbook.sheetnames or 'Counties' not in workbook.sheetnames:
            print("Raw or Counties sheet not found, skipping key standardization")
            return
        
        raw_sheet = workbook['Raw']
        counties_sheet = workbook['Counties']
        
        # Fix Raw sheet KEY column (D) - CountyYear format
        print("Standardizing Raw sheet KEY column...")
        raw_rows_fixed = 0
        row = 2  # Start from row 2 (skip header)
        
        while raw_sheet[f'A{row}'].value is not None:
            county = raw_sheet[f'A{row}'].value
            year = raw_sheet[f'C{row}'].value
            
            if county and year:
                # Set KEY to CountyYear format (e.g., "Caswell2010")
                raw_sheet[f'D{row}'] = f'=CONCAT(A{row},C{row})'
                raw_rows_fixed += 1
                
                # Also add calculated field formulas if missing
                # Column F: Deaths per 1,000
                if raw_sheet[f'F{row}'].value is None:
                    raw_sheet[f'F{row}'] = f'=IF(E{row}=0,0,(G{row}/E{row})*1000)'
                
                # Column H: Death Service Ratio
                if raw_sheet[f'H{row}'].value is None:
                    raw_sheet[f'H{row}'] = f'=IF(G{row}=0,0,I{row}/G{row})'
            
            row += 1
        
        print(f"  Fixed {raw_rows_fixed} KEY values in Raw sheet")
        
        # Fix Counties sheet KEY column (E) - CountyYear format
        print("Standardizing Counties sheet KEY column...")
        counties_rows_fixed = 0
        
        for row in range(2, counties_sheet.max_row + 1):
            county = counties_sheet[f'D{row}'].value
            year = counties_sheet[f'A{row}'].value
            
            if county and year:
                # Set KEY to CountyYear format (e.g., "Caswell2010")
                key_value = f"{county}{year}"
                counties_sheet[f'E{row}'] = key_value
                counties_rows_fixed += 1
        
        print(f"  Fixed {counties_rows_fixed} KEY values in Counties sheet")
        print(f"Successfully standardized KEY columns in both sheets")
        
    except Exception as e:
        print(f"Error standardizing KEY columns: {e}")
        # Don't fail the entire process
        pass

def restore_xlookup_formulas(counties_sheet, raw_sheet_max_row=76):
    """
    Restore XLOOKUP formulas to all data rows in the Counties sheet.
    These formulas look up data from the Raw sheet based on the Key column (E).
    
    Args:
        counties_sheet: The Counties worksheet object
        raw_sheet_max_row: Maximum row in Raw sheet (default 76)
    """
    try:
        # Define the XLOOKUP formula mappings
        # Format: counties_column -> (raw_lookup_column, raw_return_column)
        xlookup_mappings = {
            'H': 'E',   # Medicare Beneficiaries (Raw column E)
            'I': 'G',   # Medicare Deaths (Raw column G)
            'J': 'K',   # Hospice Unduplicated Beneficiaries (Raw column K)
            'K': 'I',   # Hospice Deaths (Raw column I)
            'Z': 'J',   # Hospice Penetration (Raw column J)
            'AB': 'N',  # Average Daily Census (Raw column N)
            'AC': 'M',  # Patient Days (Raw column M)
            'AD': 'L',  # Days per Patient/ALOS (Raw column L)
            'AE': 'O',  # % GIP Days (Raw column O)
            'AF': 'P',  # Average GIP Census (Raw column P)
            'AG': 'Q',  # GIP Patients (Raw column Q)
            'AH': 'R',  # Payments per Patient (Raw column R)
        }
        
        # Get the actual data range
        max_row = counties_sheet.max_row
        if max_row < 2:
            print("No data rows to update with XLOOKUP formulas")
            return
        
        # Apply XLOOKUP formulas to all data rows (starting from row 2)
        formulas_applied = 0
        for row in range(2, max_row + 1):
            # Check if this row has a key value
            key_cell = counties_sheet[f'E{row}']
            if key_cell.value:
                # Apply XLOOKUP formulas for each mapped column
                for counties_col, raw_col in xlookup_mappings.items():
                    # Build the XLOOKUP formula
                    # =XLOOKUP(E{row},Raw!$D$2:$D${raw_max},Raw!${raw_col}$2:${raw_col}${raw_max})
                    formula = f'=XLOOKUP(E{row},Raw!$D$2:$D${raw_sheet_max_row},Raw!${raw_col}$2:${raw_col}${raw_sheet_max_row})'
                    counties_sheet[f'{counties_col}{row}'] = formula
                formulas_applied += 1
        
        print(f"Successfully restored XLOOKUP formulas to {formulas_applied} rows in Counties sheet")
        
    except Exception as e:
        print(f"Error restoring XLOOKUP formulas: {e}")
        # Don't fail the entire process if formula restoration fails
        pass

def apply_counties_sheet_formatting(counties_sheet, start_row=None, end_row=None):
    """
    Apply consistent formatting to the Counties sheet with alternating row colors,
    center alignment, and proper number formatting.
    
    Args:
        counties_sheet: The Counties worksheet object
        start_row: Optional - first row to format (default: 2)
        end_row: Optional - last row to format (default: max_row)
    """
    try:
        # Define styles
        light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Determine the range to format
        if start_row is None:
            start_row = 2  # Skip header row
        if end_row is None:
            end_row = counties_sheet.max_row
        
        if end_row < start_row:
            return  # No data to format
        
        # Apply formatting to specified data rows
        for row_num in range(start_row, end_row + 1):
            # Determine fill color based on even/odd row
            if row_num % 2 == 0:
                fill_color = light_blue_fill
            else:
                fill_color = white_fill
            
            # Apply formatting to all columns A through AI (columns 1-35)
            for col_num in range(1, 36):  # A to AI is columns 1-35
                cell = counties_sheet.cell(row=row_num, column=col_num)
                
                # Apply fill color
                cell.fill = fill_color
                
                # Apply center alignment
                cell.alignment = center_alignment
                
                # Apply specific number formatting
                if col_num == 26:  # Column Z (Hospice Penetration)
                    cell.number_format = '0.00'
                elif col_num == 31:  # Column AE (% GIP Days)
                    cell.number_format = '0.00%'
                elif col_num in [8, 9, 10, 11]:  # Columns H-K (whole numbers)
                    cell.number_format = '#,##0'
                elif col_num in [28, 29, 30, 32, 33, 34, 35]:  # Columns AB-AH (various decimals)
                    if col_num == 30:  # Column AD (Days per Patient)
                        cell.number_format = '0.00'
                    elif col_num == 32:  # Column AF (Average GIP Census)
                        cell.number_format = '0.0'
                    elif col_num == 35:  # Column AH (Payments per Patient)
                        cell.number_format = '"$"#,##0'
                    else:
                        cell.number_format = '#,##0'
        
        print(f"Applied formatting to rows {start_row}-{end_row} in Counties sheet")
        
    except Exception as e:
        print(f"Error applying Counties sheet formatting: {e}")
        # Don't fail the entire process if formatting fails
        pass

def extract_county_data(county_file):
    """
    Extract data from County Trend sheet, rows 10+ (skipping header at row 9)
    Extracts columns B,C,E,G,H,I,J,K,L,M,N,O,P (12 total columns)
    Note: Column A is empty in Burke files - data starts at column B
    Stops at first empty row to extract only the PRIMARY table (first continuous data section)
    """
    debug_logger.reset_trace()
    
    try:
        # First, analyze the file
        file_info = debug_logger.log_file_info(county_file)
        
        # Check if file is a valid Excel file before attempting to load
        if not file_info.get("is_zip_file") and file_info.get("file_type") != "Old Excel format (XLS)":
            # Provide user-friendly error message
            file_size = file_info.get('file_size', 0)
            if file_size < 1000:  # Less than 1KB
                error_msg = f"File {county_file.filename} appears to be corrupted or not a valid Excel file (size: {file_size} bytes). Please re-export from Excel or ensure the file is a proper .xlsx format."
            else:
                error_msg = f"File {county_file.filename} is not a valid Excel file: {file_info.get('file_type', 'Unknown format')}"
            print(error_msg)
            debug_logger.logger.error(error_msg)
            return None
        
        # Attempt to load the workbook
        debug_logger.logger.info(f"Loading workbook: {county_file.filename}")
        county_wb = load_workbook(county_file, data_only=True)
        
        # Log sheet detection
        sheet_names = county_wb.sheetnames
        sheet_info = debug_logger.log_sheet_detection(county_wb, sheet_names)
        
        # Find County Trend sheet
        trend_sheet = None
        for sheet_name in sheet_names:
            if 'trend' in sheet_name.lower() and 'county' in sheet_name.lower():
                trend_sheet = county_wb[sheet_name]
                debug_logger.logger.info(f"Selected sheet: {sheet_name}")
                break
        
        if not trend_sheet:
            error_msg = f"No 'County Trend' sheet found in {county_file.filename}"
            print(error_msg)
            debug_logger.logger.warning(error_msg)
            return None
        
        # Extract county name from filename
        county_name = county_file.filename.replace('.xlsx', '').replace('.xls', '')
        debug_logger.logger.info(f"Processing county: {county_name}")
        
        # Log sheet dimensions
        debug_logger.logger.info(f"Sheet dimensions - Max row: {trend_sheet.max_row}, Max column: {trend_sheet.max_column}")
        
        # Extract data from rows 10+ (row 9 is headers, data starts at row 10)
        extracted_data = []
        rows_checked = 0
        stop_reason = "Unknown"
        
        for row in range(10, trend_sheet.max_row + 1):
            rows_checked += 1
            
            # Extract all required columns (CORRECT MAPPING - Column A is empty in Burke files)
            year_cell = trend_sheet[f'B{row}']                    # Year data starts at B
            medicare_enrollment_cell = trend_sheet[f'C{row}']     # Medicare Enrollment
            resident_deaths_cell = trend_sheet[f'E{row}']         # Resident Deaths (skip D)
            hospice_deaths_cell = trend_sheet[f'G{row}']          # Hospice Deaths (skip F)
            hospice_penetration_cell = trend_sheet[f'H{row}']     # Hospice Penetration
            patients_served_cell = trend_sheet[f'I{row}']         # Patients Served
            days_per_patient_cell = trend_sheet[f'J{row}']        # Days per Patient (ALOS)
            patient_days_cell = trend_sheet[f'K{row}']            # Patient Days
            avg_daily_census_cell = trend_sheet[f'L{row}']        # Average Daily Census
            gip_days_percent_cell = trend_sheet[f'M{row}']        # % GIP Days
            avg_gip_census_cell = trend_sheet[f'N{row}']          # Average GIP Census
            gip_patients_cell = trend_sheet[f'O{row}']            # GIP Patients
            payments_per_patient_cell = trend_sheet[f'P{row}']    # Payments per Patient
            
            # Log the raw values for debugging
            debug_logger.logger.debug(f"Row {row} raw values - Year: {year_cell.value}, Medicare: {medicare_enrollment_cell.value}")
            
            # Stop at first empty row (end of PRIMARY table)
            # Check if year or medicare enrollment is empty/None
            if (year_cell.value is None or year_cell.value == '' or 
                medicare_enrollment_cell.value is None or medicare_enrollment_cell.value == ''):
                stop_reason = f"Empty row detected at row {row}"
                debug_logger.log_row_extraction(row, year_cell.value, medicare_enrollment_cell.value, 
                                               "stop", "Empty year or medicare enrollment - end of PRIMARY table")
                break
            
            # Check if we have valid numeric data in this row
            if (isinstance(year_cell.value, (int, float)) and 
                isinstance(medicare_enrollment_cell.value, (int, float))):
                
                # Convert percentages to decimals where needed
                hospice_penetration_value = hospice_penetration_cell.value if hospice_penetration_cell.value else 0
                # If value is > 1, it's likely a percentage that needs to be divided by 100
                if hospice_penetration_value > 1:
                    hospice_penetration_value = hospice_penetration_value / 100
                
                gip_days_percent_value = gip_days_percent_cell.value if gip_days_percent_cell.value else 0
                # If value is > 1, it's likely a percentage that needs to be divided by 100
                if gip_days_percent_value > 1:
                    gip_days_percent_value = gip_days_percent_value / 100
                
                row_data = {
                    'county': county_name,
                    'state': 'NC',
                    'year': year_cell.value,
                    # Whole numbers (0 decimals)
                    'medicare_enrollment': int(round(medicare_enrollment_cell.value)) if medicare_enrollment_cell.value else 0,
                    'resident_deaths': round(resident_deaths_cell.value, 1) if resident_deaths_cell.value else 0,  # 1 decimal for death rates
                    'hospice_deaths': int(round(hospice_deaths_cell.value)) if hospice_deaths_cell.value else 0,
                    'patients_served': int(round(patients_served_cell.value)) if patients_served_cell.value else 0,
                    'patient_days': int(round(patient_days_cell.value)) if patient_days_cell.value else 0,
                    'gip_patients': int(round(gip_patients_cell.value)) if gip_patients_cell.value else 0,
                    # Percentages (4 decimals)
                    'hospice_penetration': round(hospice_penetration_value, 4),  # Converted to decimal
                    'gip_days_percent': round(gip_days_percent_value, 4),  # Converted to decimal
                    # 2 decimal places
                    'days_per_patient': round(days_per_patient_cell.value, 2) if days_per_patient_cell.value else 0,
                    'avg_daily_census': round(avg_daily_census_cell.value, 2) if avg_daily_census_cell.value else 0,
                    'payments_per_patient': round(payments_per_patient_cell.value, 2) if payments_per_patient_cell.value else 0,
                    # 1 decimal place
                    'avg_gip_census': round(avg_gip_census_cell.value, 1) if avg_gip_census_cell.value else 0
                }
                extracted_data.append(row_data)
                debug_logger.log_row_extraction(row, year_cell.value, medicare_enrollment_cell.value,
                                               "extract", "Valid numeric data")
            else:
                debug_logger.log_row_extraction(row, year_cell.value, medicare_enrollment_cell.value,
                                               "skip", f"Non-numeric data - Year type: {type(year_cell.value)}, Medicare type: {type(medicare_enrollment_cell.value)}")
        
        # Log extraction summary
        if rows_checked >= (trend_sheet.max_row - 9):
            stop_reason = "Reached end of sheet"
            
        summary = debug_logger.log_extraction_summary(county_name, rows_checked, len(extracted_data), stop_reason)
        
        # Save trace for analysis
        if len(extracted_data) != 15:
            # If we didn't get exactly 15 rows, save trace for debugging
            trace_file = debug_logger.save_trace_to_file(f"{county_name}_unexpected_count")
            debug_logger.logger.warning(f"Unexpected row count ({len(extracted_data)} instead of 15). Trace saved to {trace_file}")
        
        return extracted_data
        
    except zipfile.BadZipFile as e:
        error_msg = f"Bad zip file error for {county_file.filename}: {e}"
        print(error_msg)
        debug_logger.logger.error(error_msg)
        debug_logger.save_trace_to_file(f"{county_file.filename}_bad_zip")
        return None
    except Exception as e:
        error_msg = f"Error extracting data from {county_file.filename}: {e}"
        print(error_msg)
        debug_logger.logger.error(error_msg, exc_info=True)
        debug_logger.save_trace_to_file(f"{county_file.filename}_error")
        return None

def find_next_empty_row(sheet):
    """
    Find the next empty row in the Raw sheet
    """
    # Start from row 2 (assuming row 1 is headers)
    row = 2
    while sheet[f'A{row}'].value is not None:
        row += 1
    return row

def add_county_data_to_raw(raw_sheet, county_data, start_row):
    """
    Add county data to Raw sheet with proper field mapping (12 data columns + 2 calculated)
    
    Field Mapping (WITH INTENTIONAL GAPS for calculated fields):
    - Column A: County
    - Column B: State 
    - Column C: Year
    - Column D: Key (will be auto-generated by CONCAT formula)
    - Column E: Medicare Beneficiaries (from Medicare Enrollment)
    - Column F: ** CALCULATED - Deaths per 1,000 ** (DO NOT OVERWRITE)
    - Column G: Medicare Deaths (from Resident Deaths)
    - Column H: ** CALCULATED - Death Service Ratio ** (DO NOT OVERWRITE)
    - Column I: Hospice Deaths (from Hospice Deaths)
    - Column J: Hospice Penetration (from Hospice Penetration)
    - Column K: Hospice Unduplicated Beneficiaries (from Patients Served)
    - Column L: Days per Patient (from Days per Patient/ALOS)
    - Column M: Patient Days (from Patient Days)
    - Column N: Average Daily Census (from Average Daily Census)
    - Column O: % GIP Days (from % GIP Days)
    - Column P: Average GIP Census (from Average GIP Census)
    - Column Q: GIP Patients (from GIP Patients)
    - Column R: Payments per Patient (from Payments per Patient)
    """
    current_row = start_row
    
    for data in county_data:
        # Map data to Raw sheet columns
        raw_sheet[f'A{current_row}'] = data['county']
        raw_sheet[f'B{current_row}'] = data['state']
        raw_sheet[f'C{current_row}'] = data['year']
        
        # Column D: Key field - CountyYear format (e.g., "Caswell2010")
        # This must match the format used in Counties sheet column E
        raw_sheet[f'D{current_row}'] = f'=CONCAT(A{current_row},C{current_row})'
        
        # Field mapping - WITH INTENTIONAL GAPS for calculated fields
        raw_sheet[f'E{current_row}'] = data['medicare_enrollment']  # Medicare Beneficiaries
        
        # Column F: Deaths per 1,000 (CALCULATED)
        # Formula: (Medicare Deaths / Medicare Beneficiaries) * 1000
        raw_sheet[f'F{current_row}'] = f'=IF(E{current_row}=0,0,(G{current_row}/E{current_row})*1000)'
        
        raw_sheet[f'G{current_row}'] = data['resident_deaths']      # Medicare Deaths
        
        # Column H: Death Service Ratio (CALCULATED)
        # Formula: Hospice Deaths / Medicare Deaths
        raw_sheet[f'H{current_row}'] = f'=IF(G{current_row}=0,0,I{current_row}/G{current_row})'
        
        raw_sheet[f'I{current_row}'] = data['hospice_deaths']       # Hospice Deaths
        raw_sheet[f'J{current_row}'] = data['hospice_penetration']  # Hospice Penetration
        raw_sheet[f'K{current_row}'] = data['patients_served']      # Hospice Unduplicated Beneficiaries
        raw_sheet[f'L{current_row}'] = data['days_per_patient']     # Days per Patient (ALOS)
        raw_sheet[f'M{current_row}'] = data['patient_days']         # Patient Days
        raw_sheet[f'N{current_row}'] = data['avg_daily_census']     # Average Daily Census
        raw_sheet[f'O{current_row}'] = data['gip_days_percent']     # % GIP Days
        raw_sheet[f'P{current_row}'] = data['avg_gip_census']       # Average GIP Census
        raw_sheet[f'Q{current_row}'] = data['gip_patients']         # GIP Patients
        raw_sheet[f'R{current_row}'] = data['payments_per_patient'] # Payments per Patient
        
        current_row += 1
    
    return current_row
